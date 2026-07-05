from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analyzer import CompetitorProfile, MarketInsight


def _md_cell(text: str) -> str:
    return " ".join(text.split()).replace("|", "\\|")


def build_markdown(profiles: list[CompetitorProfile], insight: MarketInsight) -> str:
    lines = [
        "# Competitor Snapshot",
        "",
        f"Comparison of {len(profiles)} competitors, generated from their public websites.",
        "",
        "| Competitor | USP | Target Audience | Key Features | Pricing |",
        "|---|---|---|---|---|",
    ]
    for p in profiles:
        features = "; ".join(p.key_features)
        lines.append(
            f"| **{_md_cell(p.name)}** | {_md_cell(p.usp)} | {_md_cell(p.target_audience)} "
            f"| {_md_cell(features)} | {_md_cell(p.pricing_model)} |"
        )
    lines += ["", "## Market Analysis", "", "**Common patterns (table stakes):**", ""]
    lines += [f"- {c}" for c in insight.common_patterns] or ["- _(none)_"]
    if insight.differentiators:
        lines += ["", "**Differentiators:**", ""]
        lines += [f"- {d}" for d in insight.differentiators]
    lines += ["", "**Potential market gap:**", "", insight.market_gap or "_(none identified)_", ""]
    return "\n".join(lines)


def write_excel(profiles: list[CompetitorProfile], insight: MarketInsight, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    headers = ["Competitor", "USP", "Target Audience", "Key Features", "Pricing Model"]
    widths = [22, 40, 30, 46, 26]

    header_fill = PatternFill("solid", fgColor="1F3B57")
    header_font = Font(bold=True, color="FFFFFF")
    for col, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    for row, p in enumerate(profiles, start=2):
        values = [p.name, p.usp, p.target_audience, "\n".join(f"• {f}" for f in p.key_features), p.pricing_model]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=1).font = Font(bold=True)

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Market Insight")
    ws2.column_dimensions["A"].width = 100
    r = 1

    def section(title: str, items: list[str]) -> None:
        nonlocal r
        c = ws2.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=12)
        r += 1
        for it in items:
            cell = ws2.cell(row=r, column=1, value=f"• {it}")
            cell.alignment = Alignment(wrap_text=True)
            r += 1
        r += 1

    section("Common patterns (table stakes)", insight.common_patterns)
    if insight.differentiators:
        section("Differentiators", insight.differentiators)
    section("Potential market gap", [insight.market_gap] if insight.market_gap else [])

    wb.save(path)
